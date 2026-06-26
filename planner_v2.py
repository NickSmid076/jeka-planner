"""
planner_v2.py — RKVV JEKA Trainingsplanner v2 (volledig automatisch)
Leest DATA + LOGICA bladen, plant alle sessies automatisch in op basis van
categorie-regels. Geen handmatige voorkeuren per team — het algoritme kiest
optimaal via dag-load balancing.

Gebruik:
    python planner_v2.py --file "Trainingschema planner v2.xlsx"
"""

import argparse
import datetime
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Constanten
# ---------------------------------------------------------------------------

DAGEN = ["MA", "DI", "WO", "DO", "VR"]
DAG_LABELS = {
    "MA": "Maandag", "DI": "Dinsdag", "WO": "Woensdag",
    "DO": "Donderdag", "VR": "Vrijdag",
}
SUBVELDEN = ["Veld 6A", "Veld 6B", "Veld 7A", "Veld 7B", "Veld 1A", "Veld 1B", "Veld 2A", "Veld 2B"]
VELDEN    = ["Veld 6", "Veld 7", "Veld 1", "Veld 2"]

# Veld 2 mag alleen gebruikt worden als sessie voor 19:00 eindigt
VELD2_MAX_EIND_SLOT = 12  # slot 12 = 19:00 (= (19*60-960)/15)

START_TIJD = datetime.time(16, 0)
EIND_TIJD  = datetime.time(22, 30)
STAP_MIN   = 15

DAG_IDX = {"MA": 0, "DI": 1, "WO": 2, "DO": 3, "VR": 4}

def heeft_dag_gap(dag: str, gebruikte_dagen: set, min_gap: int = 2) -> bool:
    """True als dag minstens min_gap werkdagen verwijderd is van alle gebruikte dagen."""
    new_idx = DAG_IDX.get(dag, 99)
    return all(abs(new_idx - DAG_IDX.get(g, 99)) >= min_gap for g in gebruikte_dagen)

KLEUR_CAT = {
    "onderbouw":          "AED6F1",
    "middenbouw":         "A9DFBF",
    "bovenbouw":          "F9E79F",
    "senioren":           "F1948A",
    "senioren-selectie":  "E8A0A0",
    "bijzonder":          "D7BDE2",
}


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def make_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def time_to_slot(t: datetime.time) -> int:
    """Zet tijd om naar slot-index (16:00 = 0, stap 15 min)."""
    minuten = t.hour * 60 + t.minute
    start_min = START_TIJD.hour * 60 + START_TIJD.minute
    return (minuten - start_min) // STAP_MIN


def slot_to_time(slot: int) -> datetime.time:
    start_min = START_TIJD.hour * 60 + START_TIJD.minute
    minuten = start_min + slot * STAP_MIN
    return datetime.time(minuten // 60, minuten % 60)


def parse_tijd(s) -> datetime.time | None:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%H:%M", "%H.%M"):
        try:
            return datetime.datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Team-voorkeuren inlezen uit team_preferences.json
# ---------------------------------------------------------------------------

def _lees_team_prefs() -> dict:
    prefs_pad = Path(__file__).parent / "team_preferences.json"
    if not prefs_pad.exists():
        return {}
    try:
        with open(prefs_pad, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# DATA inlezen (versie 2)
# ---------------------------------------------------------------------------

def lees_data(wb, sheet_name="DATA") -> list[dict]:
    ws = wb[sheet_name]
    headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}

    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        team_id   = row[headers["TeamID"]]       if "TeamID"        in headers else None
        team_naam = row[headers["Team_naam"]]     if "Team_naam"     in headers else None
        categorie = row[headers["Categorie"]]     if "Categorie"     in headers else None
        actief    = row[headers["Actief"]]        if "Actief"        in headers else None
        bijz      = row[headers["Bijzonderheden"]] if "Bijzonderheden" in headers else None

        if not team_naam:
            continue
        if actief is not True and str(actief).upper() not in ("TRUE", "1"):
            continue

        bijz_str  = str(bijz).lower() if bijz else ""
        veld_m    = re.search(r"veld:(\d)",        bijz_str)
        duur_m    = re.search(r"duur:(\d+)",       bijz_str)
        gebruik_m = re.search(r"gebruik:([\d.]+)", bijz_str)
        dag_m     = re.search(r"\bdag:([a-z]{2})\b", bijz_str)

        team = {
            "team_id":        team_id or team_naam,
            "team_naam":      team_naam,
            "categorie":      categorie or "",
            "bijzonderheden": bijz_str,
            "voorkeur_dag":   dag_m.group(1).upper() if dag_m else None,
            "voorkeur_veld":  f"Veld {veld_m.group(1)}" if veld_m else None,
            "duur_override":  int(duur_m.group(1)) if duur_m else None,
            "gebruik_override": float(gebruik_m.group(1)) if gebruik_m else None,
        }
        teams.append(team)

    # Overlay met voorkeuren uit team_preferences.json
    prefs = _lees_team_prefs()
    for team in teams:
        p = prefs.get(team["team_id"], {})
        if p.get("voorkeur_dag"):
            team["voorkeur_dag"] = p["voorkeur_dag"]
        team["voorkeur_tijd"]    = parse_tijd(p["voorkeur_tijd"]) if p.get("voorkeur_tijd") else None
        team["niet_beschikbaar"] = set(p.get("niet_beschikbaar", []))

    return teams


# ---------------------------------------------------------------------------
# LOGICA inlezen
# ---------------------------------------------------------------------------

def lees_logica(wb) -> dict:
    ws = wb["LOGICA"]
    logica = {"categorie_regels": {}}

    in_cat_tabel = False
    cat_headers = []

    for row in ws.iter_rows(values_only=True):
        if row[0] and "tblCategorieRegels" in str(row[0]):
            in_cat_tabel = True
            cat_headers = []
            continue
        if in_cat_tabel:
            if not cat_headers:
                cat_headers = [str(v).strip() if v else "" for v in row]
                continue
            if row[0] is None:
                in_cat_tabel = False
                continue
            rec = dict(zip(cat_headers, row))
            cat = str(rec.get("Categorie", "")).strip()
            if cat:
                vg2_raw = rec.get("Veldgebruik_2")
                logica["categorie_regels"][cat] = {
                    "duur_min":     int(rec.get("Duur_min", 60) or 60),
                    "veldgebruik":  float(rec.get("Veldgebruik", 0.5) or 0.5),
                    "veldgebruik_2": float(vg2_raw) if vg2_raw not in (None, "") else None,
                    "sessies":      int(rec.get("Sessies_week", 2) or 2),
                    "prioriteit":   str(rec.get("Prioriteit", "")).strip(),
                    "tijd_van":     parse_tijd(rec.get("Tijd_van")),
                    "tijd_tot":     parse_tijd(rec.get("Tijd_tot")),
                }

    # Overlay met logica_overrides.json (UI-aanpassingen)
    overrides_pad = Path(__file__).parent / "logica_overrides.json"
    if overrides_pad.exists():
        try:
            with open(overrides_pad, encoding="utf-8") as _f:
                _overrides = json.load(_f)
            for _cat, _ovr in _overrides.items():
                if _cat not in logica["categorie_regels"]:
                    logica["categorie_regels"][_cat] = {
                        "duur_min": 60, "veldgebruik": 0.5, "veldgebruik_2": None,
                        "sessies": 2, "prioriteit": "bijzonder",
                        "tijd_van": datetime.time(16, 0), "tijd_tot": datetime.time(22, 30),
                    }
                _r = logica["categorie_regels"][_cat]
                if "duur_min"   in _ovr: _r["duur_min"]  = int(_ovr["duur_min"])
                if "prioriteit" in _ovr and _ovr["prioriteit"]: _r["prioriteit"] = _ovr["prioriteit"]
                if "tijd_van"   in _ovr and _ovr["tijd_van"]:  _r["tijd_van"]   = parse_tijd(_ovr["tijd_van"])
                if "tijd_tot"   in _ovr and _ovr["tijd_tot"]:  _r["tijd_tot"]   = parse_tijd(_ovr["tijd_tot"])
        except Exception:
            pass

    return logica


# ---------------------------------------------------------------------------
# Categorie-detectie op basis van TeamID-prefix
# ---------------------------------------------------------------------------

CATEGORIE_MAP = [
    (r"^JO7",  "JO7"),  (r"^JO8",  "JO8"),
    (r"^JO9",  "JO9"),  (r"^JO10", "JO10"),
    (r"^JO11", "JO11"), (r"^JO12", "JO12"),
    (r"^JO13", "JO13"), (r"^JO14", "JO14"),
    (r"^JO15", "JO15"), (r"^JO16", "JO16"), (r"^JO17", "JO17"),
    (r"^JO19", "JO19"), (r"^MO13", "MO13"),
    (r"^MO15", "MO15"), (r"^MO17", "MO17"),
    (r"^MO20", "MO20"),
    (r"^(Heren|Dames|Vaders)", "Senioren"),
    (r"^(G-|Keeperstraining|4SKILLS)", "Bijzonder"),
    (r"^Gehandicapt", "Gehandicapt"),
]


def detecteer_categorie(team_id: str) -> str:
    for pattern, cat in CATEGORIE_MAP:
        if re.match(pattern, team_id, re.IGNORECASE):
            return cat
    return "Bijzonder"


def haal_regels(team: dict, logica: dict) -> dict:
    """Haal categorie-regels op voor dit team.
    Volgorde: 1) exacte match op DATA-categorie, 2) regex op team_id, 3) fuzzy match."""
    cat_regels = logica["categorie_regels"]
    cat_data   = str(team.get("categorie", "")).strip()
    _default   = {
        "duur_min": 60, "veldgebruik": 0.5, "veldgebruik_2": None, "sessies": 2,
        "prioriteit": "bijzonder",
        "tijd_van": datetime.time(16, 0), "tijd_tot": datetime.time(22, 30),
    }

    # 1. Exacte match op categorie uit DATA (bijv. "Senioren-selectie")
    if cat_data and cat_data in cat_regels:
        regels = cat_regels[cat_data]
    else:
        # 2. Regex-detectie op team_id
        cat_key = detecteer_categorie(team["team_id"])
        if cat_key in cat_regels:
            regels = cat_regels[cat_key]
        else:
            # 3. Fuzzy: zoek cat_name die voorkomt in de DATA-categorie
            regels = next(
                (d for n, d in cat_regels.items() if n.lower() in cat_data.lower()),
                _default,
            )

    return regels


# ---------------------------------------------------------------------------
# Capaciteitsrooster
# ---------------------------------------------------------------------------

N_SLOTS = (time_to_slot(EIND_TIJD) + 1)


def leeg_grid():
    return {dag: {veld: [0.0] * N_SLOTS for veld in VELDEN} for dag in DAGEN}


def past_in_grid(grid, dag, veld, start_slot, n_slots, veldgebruik) -> bool:
    if start_slot + n_slots > N_SLOTS:
        return False
    for s in range(start_slot, start_slot + n_slots):
        if grid[dag][veld][s] + veldgebruik > 1.0 + 1e-6:
            return False
    return True


def reserveer(grid, dag, veld, start_slot, n_slots, veldgebruik):
    for s in range(start_slot, start_slot + n_slots):
        grid[dag][veld][s] += veldgebruik


# ---------------------------------------------------------------------------
# Dag-load balancing helpers
# ---------------------------------------------------------------------------

def kies_dag_volgorde(grid) -> list[str]:
    """Sorteer DAGEN op huidig veldgebruik, minst belast eerst."""
    dag_loads = {
        dag: sum(grid[dag][veld][s] for veld in VELDEN for s in range(N_SLOTS))
        for dag in DAGEN
    }
    return sorted(DAGEN, key=lambda d: dag_loads[d])


# ---------------------------------------------------------------------------
# Scorefunctie (v2 — dag-load aware, geen voorkeurstijd)
# ---------------------------------------------------------------------------

def score_v2(start_tijd: datetime.time, regels: dict, dag: str, grid: dict) -> float:
    s = 0.0
    t_van = regels.get("tijd_van") or START_TIJD
    t_tot = regels.get("tijd_tot") or EIND_TIJD

    start_min = start_tijd.hour * 60 + start_tijd.minute
    van_min   = t_van.hour * 60 + t_van.minute
    tot_min   = t_tot.hour * 60 + t_tot.minute

    # Buiten voorkeurvenster van de categorie
    if start_min < van_min:
        s += (van_min - start_min) / 60 * 10
    if start_min > tot_min:
        s += (start_min - tot_min) / 60 * 10

    # Dag-belasting penalty — voorkeur voor minst belaste dag
    dag_load = sum(grid[dag][veld][slot] for veld in VELDEN for slot in range(N_SLOTS))
    s += dag_load * 0.5

    # Lichte voorkeur voor vroeger in het tijdvenster — bovenbouw start zo bij 19:30
    # zodat het 21:00-slot vrij blijft voor senioren
    if start_min >= van_min:
        s += (start_min - van_min) / 60 * 0.15

    # Per-team voorkeurstijdstip (zachte constraint via regels doorgegeven)
    voorkeur_tijd = regels.get("voorkeur_tijd")
    if voorkeur_tijd:
        pref_min = voorkeur_tijd.hour * 60 + voorkeur_tijd.minute
        s += abs(start_min - pref_min) / 60 * 2.0

    return s


# ---------------------------------------------------------------------------
# Planningsalgoritme
# ---------------------------------------------------------------------------

def plan_sessie(team: dict, dag: str, regels: dict, grid: dict,
                geplande_sessies: list) -> dict | None:
    """Zoek het beste tijdslot voor een sessie op een dag."""
    duur_min    = regels["duur_min"]
    veldgebruik = regels["veldgebruik"]
    n_slots     = duur_min // STAP_MIN

    tijd_van = regels.get("tijd_van") or START_TIJD
    tijd_tot = regels.get("tijd_tot") or EIND_TIJD

    beste = None
    beste_score = float("inf")

    start_from = max(time_to_slot(tijd_van), 0)
    eind_slot  = time_to_slot(tijd_tot)

    for veld in VELDEN:
        for s in range(start_from, eind_slot + 1):
            if veld == "Veld 2" and (s + n_slots > VELD2_MAX_EIND_SLOT
                                       or regels.get("prioriteit") != "onderbouw"):
                continue
            start_t = slot_to_time(s)
            eind_s  = s + n_slots
            if eind_s > N_SLOTS:
                continue
            eind_t = slot_to_time(eind_s)
            if eind_t > EIND_TIJD:
                continue

            if not past_in_grid(grid, dag, veld, s, n_slots, veldgebruik):
                continue

            sc = score_v2(start_t, regels, dag, grid)

            if sc < beste_score:
                beste_score = sc
                beste = {
                    "team_id":     team["team_id"],
                    "dag":         dag,
                    "dag_label":   DAG_LABELS[dag],
                    "start":       start_t,
                    "eind":        eind_t,
                    "veld":        veld,
                    "subveld":     None,
                    "veldgebruik": veldgebruik,
                    "prioriteit":  regels["prioriteit"],
                    "n_slots":     n_slots,
                    "start_slot":  s,
                    "score":       sc,
                }

    return beste


def wijs_subvelden_toe(sessies: list[dict]):
    """Post-processing: wijs 1A/1B/1C/1D toe op basis van werkelijk bezette subvelden."""
    LABELS = ["A", "B", "C", "D"]

    # Bouw per dag/veld/slot een lijst van sessie-indices op
    slot_gebruik: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for i, ses in enumerate(sessies):
        for s in range(ses["start_slot"], ses["start_slot"] + ses["n_slots"]):
            slot_gebruik[ses["dag"]][ses["veld"]][s].append(i)

    subveld_toewijzing: dict[int, str] = {}

    # Verwerk sessies op volgorde van starttijd zodat vroege sessies eerst een label krijgen
    for i in sorted(range(len(sessies)), key=lambda x: sessies[x]["start_slot"]):
        if i in subveld_toewijzing:
            continue
        ses = sessies[i]
        veld_nr = ses["veld"][-1]

        if ses["veldgebruik"] >= 1.0:
            subveld_toewijzing[i] = f"Veld {veld_nr}A"
            continue

        # Bepaal welke subvelden al in gebruik zijn door overlappende sessies
        bezet: set[str] = set()
        for s in range(ses["start_slot"], ses["start_slot"] + ses["n_slots"]):
            for j in slot_gebruik[ses["dag"]][ses["veld"]][s]:
                if j != i and j in subveld_toewijzing:
                    bezet.add(subveld_toewijzing[j])

        # Wijs het eerste vrije label toe
        for label in LABELS:
            kandidaat = f"Veld {veld_nr}{label}"
            if kandidaat not in bezet:
                subveld_toewijzing[i] = kandidaat
                break
        else:
            subveld_toewijzing[i] = f"Veld {veld_nr}A"

    for i, ses in enumerate(sessies):
        ses["subveld"] = subveld_toewijzing.get(i, f"{ses['veld']}A")


def plan_alles(teams: list[dict], logica: dict) -> tuple[list[dict], list[dict]]:
    grid = leeg_grid()
    geplande_sessies: list[dict] = []
    niet_ingepland: list[dict] = []

    # Vaste tijdsloten en bi-wekelijkse teams eerst; schaduwteams als laatste pass
    prioriteit_1  = [t for t in teams if "vast tijdstip" in t["bijzonderheden"]
                     or "bi-wekelijks" in t["bijzonderheden"]]
    schaduw_teams = [t for t in teams if "schaduwtraining" in t["bijzonderheden"]]
    rest = [t for t in teams if t not in prioriteit_1 and t not in schaduw_teams]

    prio_volgorde = ["bijzonder", "onderbouw", "middenbouw", "senioren-selectie",
                     "bovenbouw", "senioren"]

    def plan_team(team):
        regels = haal_regels(team, logica)
        bijz   = team["bijzonderheden"]
        # Voorkeurstijdstip doorgeven aan scorefunctie via regels-dict
        if team.get("voorkeur_tijd"):
            regels = {**regels, "voorkeur_tijd": team["voorkeur_tijd"]}

        # Vast tijdstip: zet direct in grid op de opgegeven tijd en dag
        if "vast tijdstip" in bijz:
            match = re.search(r"(\d{1,2}:\d{2})", bijz)
            vaste_tijd = parse_tijd(match.group(1)) if match else None
            if vaste_tijd:
                # Dag kan in bijzonderheden staan ("vast tijdstip VR 19:00") of in voorkeur_dag
                DAG_NAMEN = {"ma": "MA", "di": "DI", "wo": "WO", "do": "DO", "vr": "VR",
                             "maandag": "MA", "dinsdag": "DI", "woensdag": "WO",
                             "donderdag": "DO", "vrijdag": "VR"}
                dag_uit_bijz = next(
                    (DAG_NAMEN[w] for w in bijz.lower().split() if w in DAG_NAMEN), None)
                voorkeur_dag = dag_uit_bijz or team.get("voorkeur_dag")
                dag_lijst    = [voorkeur_dag] if voorkeur_dag and voorkeur_dag in DAGEN else ["WO"]
                bijz_lower = bijz.lower()
                if "kwart" in bijz_lower:
                    veldgebruik = 0.25
                elif "half" in bijz_lower:
                    veldgebruik = 0.5
                else:
                    veldgebruik = (team["gebruik_override"] if team.get("gebruik_override") is not None
                                   else regels["veldgebruik"])
                n_slots = ((team["duur_override"] // STAP_MIN) if team.get("duur_override")
                           else regels["duur_min"] // STAP_MIN)
                voorkeur_veld = team.get("voorkeur_veld")
                velden_to_scan = ([voorkeur_veld] if voorkeur_veld and voorkeur_veld in VELDEN
                                  else VELDEN)
                start_slot = time_to_slot(vaste_tijd)
                for dag in dag_lijst:
                    for veld in velden_to_scan:
                        if past_in_grid(grid, dag, veld, start_slot, n_slots, veldgebruik):
                            reserveer(grid, dag, veld, start_slot, n_slots, veldgebruik)
                            geplande_sessies.append({
                                "team_id":     team["team_id"],
                                "dag":         dag,
                                "dag_label":   DAG_LABELS[dag],
                                "start":       vaste_tijd,
                                "eind":        slot_to_time(start_slot + n_slots),
                                "veld":        veld,
                                "subveld":     None,
                                "veldgebruik": veldgebruik,
                                "prioriteit":  regels["prioriteit"],
                                "n_slots":     n_slots,
                                "start_slot":  start_slot,
                                "score":       0.0,
                            })
                            break
                return

        # Schaduwteams: koppel aan het hoofdteam-slot
        if "schaduwtraining" in bijz:
            match = re.search(r"schaduwtraining\s+(\S+)", bijz)
            hoofd_id  = match.group(1).upper() if match else None
            hoofd_ses = [s for s in geplande_sessies
                         if s["team_id"].upper() == hoofd_id]
            if hoofd_ses:
                hoofd = hoofd_ses[0]
                geplande_sessies.append({
                    **hoofd,
                    "team_id": team["team_id"],
                    "subveld": None,
                    "score":   0.0,
                })
                return
            niet_ingepland.append({
                "team_id": team["team_id"],
                "reden": f"Schaduwteam: hoofdteam {hoofd_id} niet gevonden",
            })
            return

        # Gewenste sessies komen uit LOGICA (niet uit handmatige invoer)
        gewenste_sessies = 1 if "bi-wekelijks" in bijz else regels["sessies"]

        # Kies dagvolgorde op basis van huidige belasting (minst belast eerst)
        beschikbare_dagen = kies_dag_volgorde(grid)

        # Bepaal co-locatie voorkeursdagen voor team-2 (eerste sessie op zelfde veld als team-1)
        m_tnum = re.search(r"-(\d+)$", str(team.get("team_id", "")))
        co_locatie_dagen  = []
        co_locatie_slots  = []  # (dag, veld, start_slot) van team-1 sessies
        if m_tnum and int(m_tnum.group(1)) == 2:
            team1_id = re.sub(r"-2$", "-1", str(team["team_id"]))
            team1_sessies = [s for s in geplande_sessies if s["team_id"] == team1_id]
            if team1_sessies:
                co_locatie_dagen = [s["dag"] for s in team1_sessies]
                co_locatie_slots = [(s["dag"], s["veld"], s["start_slot"])
                                    for s in team1_sessies]

        sessies_gepland = 0
        gebruikte_dagen: set[str] = set()

        # Dagvolgorde: co-locatie > voorkeur_dag > load-balanced
        voorkeur_dag = team.get("voorkeur_dag")
        if co_locatie_dagen:
            dag_volgorde = (
                [d for d in co_locatie_dagen if d in beschikbare_dagen] +
                [d for d in beschikbare_dagen if d not in co_locatie_dagen]
            )
        elif voorkeur_dag and voorkeur_dag in beschikbare_dagen:
            dag_volgorde = [voorkeur_dag] + [d for d in beschikbare_dagen if d != voorkeur_dag]
        else:
            dag_volgorde = beschikbare_dagen

        # Niet-beschikbare dagen filteren (zachte constraint: fallback naar alle dagen)
        niet_beschikbaar = team.get("niet_beschikbaar", set())
        if niet_beschikbaar:
            gefilterd = [d for d in dag_volgorde if d not in niet_beschikbaar]
            dag_volgorde = gefilterd if gefilterd else dag_volgorde

        def _probeer_dag(dag: str) -> bool:
            nonlocal sessies_gepland
            if dag in gebruikte_dagen:
                return False
            r = ({**regels, "veldgebruik": regels["veldgebruik_2"]}
                 if sessies_gepland > 0 and regels.get("veldgebruik_2") is not None
                 else regels)
            ses = plan_sessie(team, dag, r, grid, geplande_sessies)
            if ses:
                reserveer(grid, dag, ses["veld"], ses["start_slot"], ses["n_slots"],
                          ses["veldgebruik"])
                geplande_sessies.append(ses)
                sessies_gepland += 1
                gebruikte_dagen.add(dag)
                return True
            return False

        # Pass 0: match exact dezelfde tijd+veld als team-1 (samen trainen)
        for (dag, veld, start_slot) in co_locatie_slots:
            if sessies_gepland >= gewenste_sessies:
                break
            if dag in gebruikte_dagen:
                continue
            r  = ({**regels, "veldgebruik": regels["veldgebruik_2"]}
                  if sessies_gepland > 0 and regels.get("veldgebruik_2") is not None
                  else regels)
            vg = r["veldgebruik"]
            ns = r["duur_min"] // STAP_MIN
            if past_in_grid(grid, dag, veld, start_slot, ns, vg):
                reserveer(grid, dag, veld, start_slot, ns, vg)
                geplande_sessies.append({
                    "team_id":     team["team_id"],
                    "dag":         dag,
                    "dag_label":   DAG_LABELS[dag],
                    "start":       slot_to_time(start_slot),
                    "eind":        slot_to_time(start_slot + ns),
                    "veld":        veld,
                    "subveld":     None,
                    "veldgebruik": vg,
                    "prioriteit":  regels["prioriteit"],
                    "n_slots":     ns,
                    "start_slot":  start_slot,
                    "score":       0.0,
                })
                sessies_gepland += 1
                gebruikte_dagen.add(dag)

        # Pass 1: probeer dagen met ≥ 2 werkdagen tussenruimte (voor 2e+ sessie)
        for dag in dag_volgorde:
            if sessies_gepland >= gewenste_sessies:
                break
            if sessies_gepland > 0 and not heeft_dag_gap(dag, gebruikte_dagen):
                continue
            _probeer_dag(dag)

        # Pass 2: fallback — plan resterende sessies zonder gap-eis
        if sessies_gepland < gewenste_sessies:
            for dag in dag_volgorde:
                if sessies_gepland >= gewenste_sessies:
                    break
                _probeer_dag(dag)

        if sessies_gepland < gewenste_sessies:
            niet_ingepland.append({
                "team_id": team["team_id"],
                "reden": (f"Slechts {sessies_gepland}/{gewenste_sessies} sessies ingepland "
                          f"— onvoldoende veldcapaciteit"),
            })

    # Plan prioriteit_1 eerst (vaste tijdsloten + bi-wekelijks)
    for team in prioriteit_1:
        plan_team(team)

    # Plan de rest gesorteerd op categorie-prioriteit
    def prio_key(t):
        regels = haal_regels(t, logica)
        prio   = regels.get("prioriteit", "bijzonder")
        return prio_volgorde.index(prio) if prio in prio_volgorde else 99

    # Plan de rest gesorteerd op categorie-prioriteit
    for team in sorted(rest, key=prio_key):
        plan_team(team)

    # Schaduwteams als laatste: hoofdteam is nu zeker al ingepland
    for team in schaduw_teams:
        plan_team(team)

    # 2e ronde: 2e sessie voor niet-selectie senioren op kwart veld (0.25)
    # Hogere teams (lager teamnummer = betere divisie) krijgen voorrang
    niet_sel_senioren = [
        t for t in rest
        if haal_regels(t, logica).get("prioriteit") == "senioren"
    ]

    def team_nummer(t):
        m = re.search(r"-(\d+)$", str(t.get("team_id", "")))
        return int(m.group(1)) if m else 99

    for team in sorted(niet_sel_senioren, key=team_nummer):
        al_gepland = [s for s in geplande_sessies if s["team_id"] == team["team_id"]]
        if not al_gepland:
            continue
        gebruikte_dagen_sen = {s["dag"] for s in al_gepland}
        regels_2e = {**haal_regels(team, logica), "veldgebruik": 0.25}
        dag_order  = kies_dag_volgorde(grid)
        gepland_2e = False
        # Pass 1: voorkeur voor dag met voldoende tussenruimte
        for dag in dag_order:
            if dag in gebruikte_dagen_sen:
                continue
            if not heeft_dag_gap(dag, gebruikte_dagen_sen):
                continue
            ses = plan_sessie(team, dag, regels_2e, grid, geplande_sessies)
            if ses:
                reserveer(grid, dag, ses["veld"], ses["start_slot"], ses["n_slots"],
                          ses["veldgebruik"])
                geplande_sessies.append(ses)
                gepland_2e = True
                break
        # Pass 2: fallback zonder gap-eis
        if not gepland_2e:
            for dag in dag_order:
                if dag in gebruikte_dagen_sen:
                    continue
                ses = plan_sessie(team, dag, regels_2e, grid, geplande_sessies)
                if ses:
                    reserveer(grid, dag, ses["veld"], ses["start_slot"], ses["n_slots"],
                              ses["veldgebruik"])
                    geplande_sessies.append(ses)
                    break

    wijs_subvelden_toe(geplande_sessies)

    return geplande_sessies, niet_ingepland


# ---------------------------------------------------------------------------
# Resultaat wegschrijven naar ROOSTER-blad
# ---------------------------------------------------------------------------

def schrijf_rooster(wb, sessies: list[dict], niet_ingepland: list[dict], sheet_name="ROOSTER"):
    ws = wb[sheet_name]

    eerste_sessierij    = ws["J1"].value or 6
    eerste_uitzondering = ws["J2"].value or 60

    if not isinstance(eerste_sessierij, int):
        eerste_sessierij = 6
    if not isinstance(eerste_uitzondering, int):
        eerste_uitzondering = 60

    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    # Wis bestaande sessie-data uit het grid
    GRID_START_ROW = 5
    for tb in range(N_SLOTS):
        for sv_idx in range(6):
            r = GRID_START_ROW + tb * 6 + sv_idx
            for dag_idx in range(5):
                col  = get_column_letter(3 + dag_idx)
                cell = ws[f"{col}{r}"]
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.fill  = make_fill("FDFEFE")

    # Vul het visuele grid
    subveld_row_offset = {sv: i for i, sv in enumerate(SUBVELDEN)}
    dag_col_offset     = {dag: i for i, dag in enumerate(DAGEN)}

    for ses in sessies:
        dag = ses["dag"]
        if dag not in dag_col_offset:
            continue
        subveld = ses.get("subveld") or ""
        if subveld not in subveld_row_offset:
            subveld = SUBVELDEN[0]

        sv_offset  = subveld_row_offset[subveld]
        dag_col    = 3 + dag_col_offset[dag]
        col_letter = get_column_letter(dag_col)

        start_slot = ses["start_slot"]
        n_slots    = ses["n_slots"]
        kleur = KLEUR_CAT.get(ses.get("prioriteit", "bijzonder"), "D7BDE2")
        tid   = ses["team_id"]

        for slot_idx in range(n_slots):
            slot = start_slot + slot_idx
            if slot >= N_SLOTS:
                break
            r    = GRID_START_ROW + slot * 6 + sv_offset
            cell = ws[f"{col_letter}{r}"]
            if isinstance(cell, MergedCell):
                continue
            cell.fill = make_fill(kleur)
            top_style    = "medium" if slot_idx == 0 else "thin"
            bottom_style = "medium" if slot_idx == n_slots - 1 else "thin"
            cell.border = Border(
                left=Side(style="medium"), right=Side(style="medium"),
                top=Side(style=top_style), bottom=Side(style=bottom_style),
            )
            cell.value = f"{tid}\n{ses['start'].strftime('%H:%M')}-{ses['eind'].strftime('%H:%M')}"
            cell.font  = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

    # Sessietabel
    for r in range(eerste_sessierij, eerste_sessierij + 300):
        for col in range(1, 8):
            cell = ws[f"{get_column_letter(col)}{r}"]
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, ses in enumerate(sessies):
        r   = eerste_sessierij + i
        rij = [
            ses["team_id"],
            ses["dag_label"],
            ses["start"].strftime("%H:%M"),
            ses["eind"].strftime("%H:%M"),
            ses["veld"],
            ses.get("subveld", ""),
            "Ingepland",
        ]
        for col_idx, val in enumerate(rij, 1):
            cell = ws[f"{get_column_letter(col_idx)}{r}"]
            if isinstance(cell, MergedCell):
                continue
            cell.value     = val
            cell.border    = make_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 7:
                cell.font = Font(color="27AE60", bold=True)

    # Uitzonderingsrapport
    for r in range(eerste_uitzondering, eerste_uitzondering + 200):
        for col in range(1, 3):
            cell = ws[f"{get_column_letter(col)}{r}"]
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, item in enumerate(niet_ingepland):
        r = eerste_uitzondering + i
        ws[f"A{r}"].value  = item["team_id"]
        ws[f"B{r}"].value  = item["reden"]
        ws[f"A{r}"].font   = Font(color="C0392B", bold=True)
        ws[f"B{r}"].font   = Font(color="C0392B")
        ws[f"A{r}"].border = make_border()
        ws[f"B{r}"].border = make_border()

    ws["H1"] = f"Gegenereerd: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["H1"].font      = Font(italic=True, size=9, color="FFFFFF")
    ws["H1"].alignment = Alignment(horizontal="right", vertical="center")

    print(f"  {len(sessies)} sessies weggeschreven naar {sheet_name}-blad.")
    print(f"  {len(niet_ingepland)} teams niet ingepland.")


# ---------------------------------------------------------------------------
# Rapportage
# ---------------------------------------------------------------------------

def druk_rapport(sessies, niet_ingepland):
    print("\n" + "=" * 60)
    print("PLANNINGSRESULTAAT (versie 2 — volledig automatisch)")
    print("=" * 60)

    per_dag = defaultdict(list)
    for ses in sessies:
        per_dag[ses["dag"]].append(ses)

    for dag in DAGEN:
        if dag not in per_dag:
            continue
        print(f"\n{DAG_LABELS[dag].upper()} ({len(per_dag[dag])} sessies):")
        for ses in sorted(per_dag[dag], key=lambda s: s["start"]):
            print(f"  {ses['start'].strftime('%H:%M')}-{ses['eind'].strftime('%H:%M')} "
                  f"| {ses['veld']} ({ses.get('subveld','')}) "
                  f"| {ses['team_id']}")

    if niet_ingepland:
        print("\nNIET INGEPLAND:")
        for item in niet_ingepland:
            print(f"  ✗ {item['team_id']} — {item['reden']}")

    sessies_per_dag = {dag: len(per_dag[dag]) for dag in DAGEN if dag in per_dag}
    print(f"\nTotaal ingepland: {len(sessies)} sessies")
    print(f"Niet ingepland:   {len(niet_ingepland)} teams")
    print(f"Spreiding per dag: {sessies_per_dag}")


# ---------------------------------------------------------------------------
# JSON export voor dashboard
# ---------------------------------------------------------------------------

def exporteer_json(wb_pad: str, sessies: list, niet_ingepland: list, logica: dict):
    import json
    import os

    export = {
        "generated_at": datetime.datetime.now().isoformat(),
        "sessies": [
            {
                "team_id":     s["team_id"],
                "dag":         s["dag"],
                "dag_label":   s.get("dag_label", s["dag"]),
                "start":       s["start"].strftime("%H:%M"),
                "eind":        s["eind"].strftime("%H:%M"),
                "veld":        s["veld"],
                "subveld":     s.get("subveld") or "",
                "veldgebruik": s["veldgebruik"],
                "prioriteit":  s.get("prioriteit", "bijzonder"),
            }
            for s in sessies
        ],
        "niet_ingepland": [
            {"team_id": n["team_id"], "reden": n.get("reden", "")}
            for n in niet_ingepland
        ],
        "categorie_regels": {
            cat: {
                "duur_min":    r["duur_min"],
                "veldgebruik": r["veldgebruik"],
                "sessies":     r["sessies"],
                "prioriteit":  r["prioriteit"],
                "tijd_van":    r["tijd_van"].strftime("%H:%M") if r.get("tijd_van") else "16:00",
                "tijd_tot":    r["tijd_tot"].strftime("%H:%M") if r.get("tijd_tot") else "22:30",
            }
            for cat, r in logica["categorie_regels"].items()
        },
        "stats": {
            "totaal_ingepland":      len(sessies),
            "totaal_niet_ingepland": len(niet_ingepland),
        },
    }

    json_pad = os.path.join(os.path.dirname(os.path.abspath(wb_pad)), "roster_export.json")
    with open(json_pad, "w", encoding="utf-8") as f:
        json.dump(export, f, ensure_ascii=False, indent=2)
    print(f"   JSON geëxporteerd: {json_pad}")


# ---------------------------------------------------------------------------
# Hoofdprogramma
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="RKVV JEKA Trainingsplanner v2 — volledig automatisch op basis van categorieregels"
    )
    parser.add_argument("--file", default="Trainingschema planner v2.xlsx",
                        help="Pad naar het Excel-bestand")
    parser.add_argument("--data-sheet",   default="DATA",
                        help="Naam van het DATA-blad (default: DATA)")
    parser.add_argument("--rooster-sheet", default="ROOSTER",
                        help="Naam van het ROOSTER-blad (default: ROOSTER)")
    args = parser.parse_args()

    print(f"Planner v2 gestart — bestand: {args.file}")
    if args.data_sheet != "DATA" or args.rooster_sheet != "ROOSTER":
        print(f"  Bladen: {args.data_sheet} → {args.rooster_sheet}")

    wb = openpyxl.load_workbook(args.file)

    print("1. DATA inlezen...")
    teams = lees_data(wb, args.data_sheet)
    print(f"   {len(teams)} actieve teams gevonden.")

    print("2. LOGICA inlezen...")
    logica = lees_logica(wb)
    print(f"   {len(logica['categorie_regels'])} categorieregels geladen.")

    print("3. Planning uitvoeren (volledig automatisch)...")
    sessies, niet_ingepland = plan_alles(teams, logica)

    druk_rapport(sessies, niet_ingepland)

    print(f"\n4. Resultaat wegschrijven naar {args.rooster_sheet}-blad...")
    schrijf_rooster(wb, sessies, niet_ingepland, args.rooster_sheet)

    wb.save(args.file)
    print(f"   Opgeslagen: {args.file}")

    print("\n5. JSON exporteren voor dashboard...")
    exporteer_json(args.file, sessies, niet_ingepland, logica)


if __name__ == "__main__":
    main()
