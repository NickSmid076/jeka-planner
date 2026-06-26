#!/usr/bin/env python3
"""
JEKA Capaciteitsmanager Server
Start met: python3 capacity_server.py
Opent automatisch http://localhost:5051
"""

import importlib.util
import io
import json
import re
import subprocess
import sys
import webbrowser
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from threading import Timer

import openpyxl

BASE        = Path(__file__).parent
APP_DIR     = BASE / "capacity_app"
ROSTER      = BASE / "roster_export.json"
EXCEL       = BASE / "Trainingschema planner v2.xlsx"
PREFS_FILE       = BASE / "team_preferences.json"
LOGICA_OVERRIDES = BASE / "logica_overrides.json"
PLANNER_PY  = BASE / "planner_v2.py"
SEASONS_DIR = BASE / "seasons"

MIME = {".html": "text/html", ".css": "text/css", ".js": "application/javascript",
        ".json": "application/json", ".ico": "image/x-icon"}

# ── Seizoenenmap aanmaken en bestaand rooster migreren ────────────
SEASONS_DIR.mkdir(exist_ok=True)
_DEFAULT_SLUG  = "2025_2026"
_DEFAULT_LABEL = "2025/2026"

def _migreer_roster():
    dst = SEASONS_DIR / f"{_DEFAULT_SLUG}.json"
    if not dst.exists() and ROSTER.exists():
        with ROSTER.open(encoding="utf-8") as f:
            data = json.load(f)
        data.setdefault("seizoen", _DEFAULT_LABEL)
        with dst.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

_migreer_roster()

# ── Planner importeren als module ─────────────────────────────────
_planner = None
if PLANNER_PY.exists():
    try:
        _pspec = importlib.util.spec_from_file_location("planner_v2", PLANNER_PY)
        _planner = importlib.util.module_from_spec(_pspec)
        _pspec.loader.exec_module(_planner)
    except Exception as e:
        print(f"[warn] Planner kon niet worden geladen: {e}")

# ── Teams genereren vanuit seizoensspecificatie ───────────────────

# Senioren-types: (prefix, categorie_selectie, categorie_regulier)
_SENIOR_CATS = {
    "Heren":  ("Heren",  "Senioren-selectie", "Senioren"),
    "Dames":  ("Dames",  "Senioren-selectie", "Senioren"),
    "Vaders": ("Vaders", "Senioren",          "Senioren"),  # geen selectie-onderscheid
}

# Mapping formulier-categorieën → LOGICA-categorie (voor categorieën niet in LOGICA)
_LOGICA_CAT = {
    # Jongens — rechtstreeks in LOGICA
    "JO7": "JO7", "JO8": "JO8", "JO9": "JO9", "JO10": "JO10", "JO11": "JO11",
    "JO12": "JO12", "JO13": "JO13", "JO14": "JO14", "JO15": "JO15",
    "JO16": "JO17", "JO17": "JO17",
    "JO18": "JO19", "JO19": "JO19",
    # Jongens bovenbouw extra (gebruik senioren-regels)
    "JO21": "Senioren", "JO23": "Senioren",
    # Meisjes onderbouw (gebruik JO-equivalent)
    "MO8": "JO8", "MO9": "JO9", "MO10": "JO10", "MO11": "JO11",
    # Meisjes middenbouw
    "MO12": "JO12", "MO13": "MO13", "MO14": "JO14", "MO15": "MO15",
    # Meisjes bovenbouw
    "MO17": "MO17", "MO19": "JO19", "MO20": "MO20",
    # Overig
    "G-teams":        "Gehandicapt",   # gemengd G-voetbal
    "Keeperstraining": "Bijzonder",
}

# 4SKILLS: 3 vaste trainingen op vrijdag, Veld 1 — worden als aparte teams aangemaakt
_4SKILLS_TEAMS = [
    {   # VR 16:00–17:00, heel Veld 1 (1A + 1B)
        "team_id": "4SKILLS-1", "team_naam": "4SKILLS 1",
        "categorie": "Bijzonder", "actief": True,
        "bijzonderheden": "vast tijdstip VR 16:00",
        "voorkeur_dag": "VR", "voorkeur_veld": "Veld 1",
        "duur_override": 60, "gebruik_override": 1.0,
    },
    {   # VR 17:00–18:15, alleen Veld 1A
        "team_id": "4SKILLS-2", "team_naam": "4SKILLS 2",
        "categorie": "Bijzonder", "actief": True,
        "bijzonderheden": "vast tijdstip VR 17:00",
        "voorkeur_dag": "VR", "voorkeur_veld": "Veld 1",
        "duur_override": 75, "gebruik_override": None,
    },
    {   # VR 18:15–19:30, alleen Veld 1A
        "team_id": "4SKILLS-3", "team_naam": "4SKILLS 3",
        "categorie": "Bijzonder", "actief": True,
        "bijzonderheden": "vast tijdstip VR 18:15",
        "voorkeur_dag": "VR", "voorkeur_veld": "Veld 1",
        "duur_override": 75, "gebruik_override": None,
    },
]

def genereer_teams_uit_spec(spec: dict) -> list:
    teams = []
    for cat_key, aantallen in spec.items():
        totaal   = max(0, int(aantallen.get("totaal", 0)))
        selectie = min(totaal, max(0, int(aantallen.get("selectie", 0))))

        if cat_key == "4SKILLS":
            if totaal > 0:
                teams.extend(_4SKILLS_TEAMS)
        elif cat_key in _SENIOR_CATS:
            prefix, cat_sel, cat_reg = _SENIOR_CATS[cat_key]
            for i in range(1, totaal + 1):
                is_sel = i <= selectie
                team = {
                    "team_id":          f"{prefix}-{i}",
                    "team_naam":        f"{prefix} {i}",
                    "categorie":        cat_sel if is_sel else cat_reg,
                    "actief":           True,
                    "bijzonderheden":   "",
                    "voorkeur_dag":     None,
                    "voorkeur_veld":    None,
                    "duur_override":    None,
                    "gebruik_override": None,
                }
                # Heren selectie traint op dinsdag én donderdag
                if cat_key == "Heren" and is_sel:
                    team["voorkeur_dagen"] = ["DI", "DO"]
                teams.append(team)
        else:
            logica_cat = _LOGICA_CAT.get(cat_key, cat_key)
            for i in range(1, totaal + 1):
                teams.append({
                    "team_id":          f"{cat_key}-{i}",
                    "team_naam":        f"{cat_key} {i}",
                    "categorie":        logica_cat,
                    "actief":           True,
                    "bijzonderheden":   "selectie" if i <= selectie else "",
                    "voorkeur_dag":     None,
                    "voorkeur_veld":    None,
                    "duur_override":    None,
                    "gebruik_override": None,
                })
    return teams

def _time_to_str(t):
    import datetime as _dt
    if isinstance(t, _dt.time):
        return t.strftime("%H:%M")
    return str(t)

def _seizoen_slug(seizoen: str) -> str:
    return seizoen.replace("/", "_")

def _lees_seizoenen() -> list:
    seizoenen = []
    for p in sorted(SEASONS_DIR.glob("*.json")):
        try:
            with p.open(encoding="utf-8") as f:
                d = json.load(f)
            seizoenen.append({
                "slug":         p.stem,
                "seizoen":      d.get("seizoen", p.stem.replace("_", "/")),
                "generated_at": d.get("generated_at", ""),
            })
        except Exception:
            pass
    return seizoenen


def lees_teams_excel():
    """Lees alle teams uit het Excel DATA-blad (inclusief inactieve)."""
    if not EXCEL.exists():
        return {"teams": [], "error": "Excel bestand niet gevonden"}
    try:
        wb = openpyxl.load_workbook(str(EXCEL), read_only=True, data_only=True)
        ws = wb["DATA"]
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.value] = cell.column - 1
        teams = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            team_id   = row[headers.get("TeamID", 0)]
            team_naam = row[headers.get("Team_naam", 1)]
            categorie = row[headers.get("Categorie", 2)]
            actief    = row[headers.get("Actief", 3)]
            bijz      = row[headers.get("Bijzonderheden", 4)]
            if not team_naam:
                continue
            is_actief = actief is True or str(actief).upper() in ("TRUE", "1")
            teams.append({
                "team_id":        str(team_id) if team_id else str(team_naam),
                "team_naam":      str(team_naam),
                "categorie":      str(categorie) if categorie else "",
                "actief":         is_actief,
                "bijzonderheden": str(bijz) if bijz else "",
            })
        wb.close()
        return {"teams": teams}
    except Exception as e:
        return {"teams": [], "error": str(e)}


def zet_team_actief(team_id, actief):
    """Zet de Actief-waarde van een team in het Excel DATA-blad."""
    if not EXCEL.exists():
        return {"ok": False, "error": "Excel bestand niet gevonden"}
    try:
        wb = openpyxl.load_workbook(str(EXCEL))
        ws = wb["DATA"]
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        actief_col = headers.get("Actief")
        team_id_col = headers.get("TeamID")
        team_naam_col = headers.get("Team_naam")
        if not actief_col:
            return {"ok": False, "error": "Kolom 'Actief' niet gevonden"}
        gevonden = False
        for row in ws.iter_rows(min_row=2):
            id_val   = row[team_id_col - 1].value   if team_id_col else None
            naam_val = row[team_naam_col - 1].value if team_naam_col else None
            if str(id_val) == team_id or str(naam_val) == team_id:
                row[actief_col - 1].value = actief
                gevonden = True
        if not gevonden:
            return {"ok": False, "error": f"Team '{team_id}' niet gevonden"}
        wb.save(str(EXCEL))
        wb.close()
        return {"ok": True, "team_id": team_id, "actief": actief}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _genereer_excel(data: dict) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DAGEN      = ["MA", "DI", "WO", "DO", "VR"]
    DAG_LABELS = {"MA": "Maandag", "DI": "Dinsdag", "WO": "Woensdag",
                  "DO": "Donderdag", "VR": "Vrijdag"}
    SUBVELDEN  = ["Veld 6A", "Veld 6B", "Veld 7A", "Veld 7B",
                  "Veld 1A", "Veld 1B", "Veld 2A", "Veld 2B"]
    SUB_LABELS = ["V6A", "V6B", "V7A", "V7B", "V1A", "V1B", "V2A", "V2B"]
    DAG_FILL   = ["E8F5E9", "E3F2FD", "FFF8E1", "F3E5F5", "FBE9E7"]

    KLEUR_CAT = {
        "onderbouw":         "AED6F1",
        "middenbouw":        "A9DFBF",
        "bovenbouw":         "F9E79F",
        "senioren":          "F1948A",
        "senioren-selectie": "E8A0A0",
        "bijzonder":         "D7BDE2",
    }
    TEXT_CAT = {
        "onderbouw":         "154360",
        "middenbouw":        "145A32",
        "bovenbouw":         "7D6608",
        "senioren":          "641E16",
        "senioren-selectie": "641E16",
        "bijzonder":         "4A235A",
    }

    START_MIN = 960   # 16:00
    STAP_MIN  = 15
    N_SLOTS   = 26    # slot 0=16:00, slot 25=22:15 (eindigt 22:30)
    N_SUB     = len(SUBVELDEN)
    DATA_ROW  = 3     # rij 1=dag-headers, rij 2=subveld-labels, rij 3+ = tijdslots

    def dag_col(dag_idx, sub_idx):
        return 2 + dag_idx * N_SUB + sub_idx

    def slot_tijd(slot):
        m = START_MIN + slot * STAP_MIN
        return f"{m // 60:02d}:{m % 60:02d}"

    def parse_t(s):
        h, m = map(int, s.split(":"))
        return h * 60 + m

    thin   = Side(style="thin",   color="DDDDDD")
    medium = Side(style="medium", color="AAAAAA")
    no_b   = Side(style=None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekrooster"

    # Kolombreedte
    ws.column_dimensions["A"].width = 6
    for di in range(5):
        for si in range(N_SUB):
            ws.column_dimensions[get_column_letter(dag_col(di, si))].width = 11

    # Rijhoogtes
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    for sl in range(N_SLOTS):
        ws.row_dimensions[DATA_ROW + sl].height = 16

    # Rij 1: linkerbovencel
    c = ws.cell(row=1, column=1, value="Tijd")
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1A252F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=thin, right=medium, top=thin, bottom=thin)

    # Rij 1: dag-headers (samengevoegd per dag)
    for di, dag in enumerate(DAGEN):
        sc = dag_col(di, 0)
        ec = dag_col(di, N_SUB - 1)
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        c = ws.cell(row=1, column=sc, value=DAG_LABELS[dag])
        c.font      = Font(bold=True, size=12, color="1A252F")
        c.fill      = PatternFill("solid", fgColor=DAG_FILL[di])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(left=medium, right=medium, top=thin, bottom=thin)

    # Rij 2: subveld-labels
    c = ws.cell(row=2, column=1)
    c.fill   = PatternFill("solid", fgColor="1A252F")
    c.border = Border(left=thin, right=medium, top=thin, bottom=thin)

    for di in range(5):
        fill = PatternFill("solid", fgColor=DAG_FILL[di])
        for si, label in enumerate(SUB_LABELS):
            col = dag_col(di, si)
            c = ws.cell(row=2, column=col, value=label)
            c.font      = Font(bold=True, size=8, color="555555")
            c.fill      = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = Border(
                left=medium if si == 0 else thin,
                right=medium if si == N_SUB - 1 else thin,
                top=thin, bottom=thin,
            )

    # Datarijen: tijdlabels + lege achtergrond
    for sl in range(N_SLOTS):
        row     = DATA_ROW + sl
        is_hour = (sl % 4 == 0)
        is_half = (sl % 4 == 2)

        c = ws.cell(row=row, column=1,
                    value=slot_tijd(sl) if (is_hour or is_half) else None)
        c.font      = Font(size=7 if is_half else 8, bold=is_hour, color="777777")
        c.fill      = PatternFill("solid", fgColor="EAECEE" if is_hour else "F8F9FA")
        c.alignment = Alignment(horizontal="right", vertical="top")
        c.border    = Border(right=medium,
                             top=Side(style="medium" if is_hour else "thin",
                                      color="BBBBBB" if is_hour else "DDDDDD"),
                             bottom=no_b, left=no_b)

        for di in range(5):
            bg = "F0F0F0" if is_hour else "FAFAFA"
            fill = PatternFill("solid", fgColor=bg)
            for si in range(N_SUB):
                col = dag_col(di, si)
                c = ws.cell(row=row, column=col)
                c.fill   = fill
                c.border = Border(
                    left=medium if si == 0 else thin,
                    right=medium if si == N_SUB - 1 else thin,
                    top=Side(style="medium" if is_hour else "thin",
                             color="CCCCCC" if is_hour else "EEEEEE"),
                    bottom=no_b,
                )

    # Sessies plaatsen
    dag_map = {d: i for i, d in enumerate(DAGEN)}
    sub_map = {sv: i for i, sv in enumerate(SUBVELDEN)}
    # Fallback: hoofdveld → eerste subveld
    veld_fb = {}
    for sv in SUBVELDEN:
        hv = " ".join(sv.split()[:-1])  # "Veld 6A" → "Veld 6"
        if hv not in veld_fb:
            veld_fb[hv] = sv

    for s in data.get("sessies", []):
        dag = s.get("dag", "")
        if dag not in dag_map:
            continue
        sv = s.get("subveld") or veld_fb.get(s.get("veld", ""), "")
        if sv not in sub_map:
            continue

        start_sl = (parse_t(s["start"]) - START_MIN) // STAP_MIN
        n_sl     = (parse_t(s["eind"])  - parse_t(s["start"])) // STAP_MIN
        if start_sl < 0 or start_sl >= N_SLOTS or n_sl <= 0:
            continue
        n_sl = min(n_sl, N_SLOTS - start_sl)

        col   = dag_col(dag_map[dag], sub_map[sv])
        row_s = DATA_ROW + start_sl
        row_e = row_s + n_sl - 1
        prio  = s.get("prioriteit", "bijzonder")
        si    = sub_map[sv]

        if n_sl > 1:
            ws.merge_cells(start_row=row_s, start_column=col,
                           end_row=row_e,   end_column=col)

        c = ws.cell(row=row_s, column=col)
        c.value     = f"{s['team_id']}\n{s['start']}–{s['eind']}"
        c.fill      = PatternFill("solid", fgColor=KLEUR_CAT.get(prio, "D7BDE2"))
        c.font      = Font(bold=True, size=8, color=TEXT_CAT.get(prio, "4A235A"))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(
            left=medium if si == 0 else Side(style="thin", color="999999"),
            right=medium if si == N_SUB - 1 else Side(style="thin", color="999999"),
            top=medium, bottom=medium,
        )

    ws.freeze_panes = "B3"

    # Sheet 2: Niet ingepland
    ws2 = wb.create_sheet("Niet ingepland")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 45
    ws2.cell(row=1, column=1, value="Team").font  = Font(bold=True)
    ws2.cell(row=1, column=2, value="Reden").font = Font(bold=True)
    for r, n in enumerate(data.get("niet_ingepland", []), start=2):
        c1 = ws2.cell(row=r, column=1, value=n.get("team_id", ""))
        c1.font = Font(color="C0392B", bold=True)
        ws2.cell(row=r, column=2, value=n.get("reden", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        path = self.path.split("?")[0]

        if path in ("/", "/index.html"):
            self._file(APP_DIR / "index.html", "text/html")

        elif path == "/api/roster":
            self._file(ROSTER, "application/json")

        elif path == "/api/teams":
            self._json(lees_teams_excel())

        elif path == "/api/team-preferences":
            prefs = {}
            if PREFS_FILE.exists():
                with PREFS_FILE.open(encoding="utf-8") as f:
                    prefs = json.load(f)
            self._json(prefs)

        elif path == "/api/logica-regels":
            # Basisregels uit roster_export.json, overlaid met logica_overrides.json
            base = {}
            if ROSTER.exists():
                try:
                    with ROSTER.open(encoding="utf-8") as f:
                        base = json.load(f).get("categorie_regels", {})
                except Exception:
                    pass
            overrides = {}
            if LOGICA_OVERRIDES.exists():
                try:
                    with LOGICA_OVERRIDES.open(encoding="utf-8") as f:
                        overrides = json.load(f)
                except Exception:
                    pass
            merged = {}
            for cat, regels in base.items():
                merged[cat] = {**regels, **(overrides.get(cat) or {})}
            self._json(merged)

        elif path == "/api/seasons":
            self._json({"seasons": _lees_seizoenen()})

        elif re.match(r"^/api/seasons/[\w]+/roster$", path):
            slug = path.split("/")[3]
            fp   = SEASONS_DIR / f"{slug}.json"
            if fp.exists():
                self._file(fp, "application/json")
            else:
                self.send_error(404, f"Seizoen '{slug}' niet gevonden")

        elif re.match(r"^/api/seasons/[\w]+/export-excel$", path):
            slug = path.split("/")[3]
            fp   = SEASONS_DIR / f"{slug}.json"
            if not fp.exists():
                self.send_error(404, f"Seizoen '{slug}' niet gevonden")
                return
            try:
                with fp.open(encoding="utf-8") as f:
                    data = json.load(f)
                xlsx_bytes = _genereer_excel(data)
                self.send_response(200)
                self.send_header("Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition",
                    f'attachment; filename="rooster_{slug}.xlsx"')
                self.send_header("Content-Length", str(len(xlsx_bytes)))
                self._cors()
                self.end_headers()
                self.wfile.write(xlsx_bytes)
            except Exception as e:
                import traceback; traceback.print_exc()
                self.send_error(500, str(e))

        else:
            fp  = APP_DIR / path.lstrip("/")
            ext = fp.suffix.lower()
            if fp.is_file():
                self._file(fp, MIME.get(ext, "application/octet-stream"))
            else:
                self.send_error(404)

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body   = self.rfile.read(length)

        if self.path == "/api/refresh":
            try:
                # Sla huidig rooster op voor fallback (teams die nu ingepland zijn)
                oude_sessies = {}
                if ROSTER.exists():
                    with ROSTER.open(encoding="utf-8") as f:
                        oud = json.load(f)
                    for s in oud.get("sessies", []):
                        oude_sessies.setdefault(s["team_id"], []).append(s)

                r = subprocess.run(
                    [sys.executable, str(BASE / "planner_v2.py"), "--file", str(EXCEL)],
                    capture_output=True, text=True, cwd=str(BASE),
                )

                # Fallback: teams die voorheen ingepland waren maar nu niet meer → herstel
                if r.returncode == 0 and ROSTER.exists() and oude_sessies:
                    with ROSTER.open(encoding="utf-8") as f:
                        nieuw = json.load(f)
                    niet_ids = {n["team_id"] for n in nieuw.get("niet_ingepland", [])}
                    herstel  = []
                    for team_id in list(niet_ids):
                        if team_id in oude_sessies:
                            herstel.extend(oude_sessies[team_id])
                            nieuw["niet_ingepland"] = [
                                n for n in nieuw["niet_ingepland"] if n["team_id"] != team_id
                            ]
                    if herstel:
                        nieuw["sessies"].extend(herstel)
                        with ROSTER.open("w", encoding="utf-8") as f:
                            json.dump(nieuw, f, ensure_ascii=False, indent=2)

                self._json({
                    "ok":     r.returncode == 0,
                    "output": r.stdout[-3000:] if r.stdout else "",
                    "error":  r.stderr[-500:]  if r.stderr else "",
                })
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif self.path == "/api/team-preferences":
            try:
                payload  = json.loads(body)
                team_id  = payload["team_id"]
                prefs    = {}
                if PREFS_FILE.exists():
                    with PREFS_FILE.open(encoding="utf-8") as f:
                        prefs = json.load(f)
                prefs[team_id] = {
                    "voorkeur_dag":  payload.get("voorkeur_dag") or None,
                    "voorkeur_tijd": payload.get("voorkeur_tijd") or None,
                    "niet_beschikbaar": payload.get("niet_beschikbaar", []),
                }
                with PREFS_FILE.open("w", encoding="utf-8") as f:
                    json.dump(prefs, f, ensure_ascii=False, indent=2)
                self._json({"ok": True, "team_id": team_id})
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif self.path == "/api/logica-regels":
            try:
                payload = json.loads(body)
                cat     = payload["cat"]
                ovrs    = {}
                if LOGICA_OVERRIDES.exists():
                    with LOGICA_OVERRIDES.open(encoding="utf-8") as f:
                        ovrs = json.load(f)
                entry = ovrs.get(cat, {})
                if "duur_min"  in payload: entry["duur_min"]  = int(payload["duur_min"])
                if "tijd_van"  in payload: entry["tijd_van"]  = payload["tijd_van"]
                if "tijd_tot"  in payload: entry["tijd_tot"]  = payload["tijd_tot"]
                ovrs[cat] = entry
                with LOGICA_OVERRIDES.open("w", encoding="utf-8") as f:
                    json.dump(ovrs, f, ensure_ascii=False, indent=2)
                self._json({"ok": True, "cat": cat})
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif self.path == "/api/teams/toggle":
            try:
                payload  = json.loads(body)
                team_id  = payload["team_id"]
                actief   = bool(payload["actief"])
                result   = zet_team_actief(team_id, actief)
                self._json(result)
            except (KeyError, json.JSONDecodeError) as e:
                self.send_error(400, str(e))
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif self.path == "/api/roster/save":
            try:
                payload = json.loads(body)
                sessies = payload.get("sessies", [])

                existing = {}
                if ROSTER.exists():
                    with ROSTER.open(encoding="utf-8") as f:
                        existing = json.load(f)
                existing["sessies"]      = sessies
                existing["generated_at"] = datetime.now(timezone.utc).isoformat()
                with ROSTER.open("w", encoding="utf-8") as f:
                    json.dump(existing, f, ensure_ascii=False, indent=2)

                if PLANNER_PY.exists():
                    subprocess.run(
                        [sys.executable, str(PLANNER_PY),
                         "--file", str(EXCEL),
                         "--from-json", str(ROSTER)],
                        timeout=30,
                        cwd=str(PLANNER_PY.parent),
                        check=False,
                    )
                self._json({"ok": True, "count": len(sessies)})
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif self.path == "/api/seasons/create":
            try:
                payload = json.loads(body)
                seizoen = payload.get("seizoen", "").strip()

                if not re.match(r"^\d{4}/\d{4}$", seizoen):
                    self._json({"ok": False, "detail": "Seizoen moet het formaat YYYY/YYYY hebben"})
                    return

                slug    = _seizoen_slug(seizoen)
                out_fp  = SEASONS_DIR / f"{slug}.json"
                if out_fp.exists():
                    self._json({"ok": False, "detail": f"Seizoen '{seizoen}' bestaat al"})
                    return

                if not _planner:
                    self._json({"ok": False, "detail": "Planner niet beschikbaar op de server"})
                    return

                spec  = payload.get("teams", {})
                teams = genereer_teams_uit_spec(spec)
                if not teams:
                    self._json({"ok": False, "detail": "Geen teams opgegeven"})
                    return

                wb              = openpyxl.load_workbook(str(EXCEL), read_only=True, data_only=True)
                logica          = _planner.lees_logica(wb)
                wb.close()
                sessies, niet   = _planner.plan_alles(teams, logica)

                sessies_json = [
                    {
                        "team_id":    s["team_id"],
                        "dag":        s["dag"],
                        "dag_label":  s["dag_label"],
                        "start":      _time_to_str(s["start"]),
                        "eind":       _time_to_str(s["eind"]),
                        "veld":       s["veld"],
                        "subveld":    s["subveld"],
                        "veldgebruik": s["veldgebruik"],
                        "prioriteit": s["prioriteit"],
                    }
                    for s in sessies
                ]

                data = {
                    "seizoen":        seizoen,
                    "generated_at":   datetime.now(timezone.utc).isoformat(),
                    "sessies":        sessies_json,
                    "niet_ingepland": niet,
                    "stats": {
                        "totaal_ingepland":      len(sessies_json),
                        "totaal_niet_ingepland": len(niet),
                    },
                }

                with out_fp.open("w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)

                self._json({
                    "ok":                   True,
                    "slug":                 slug,
                    "seizoen":              seizoen,
                    "totaal_ingepland":     len(sessies_json),
                    "niet_ingepland_count": len(niet),
                })
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif re.match(r"^/api/seasons/[\w]+/roster/save$", self.path):
            try:
                slug    = self.path.split("/")[3]
                fp      = SEASONS_DIR / f"{slug}.json"
                payload = json.loads(body)
                sessies = payload.get("sessies", [])

                existing = {}
                if fp.exists():
                    with fp.open(encoding="utf-8") as f:
                        existing = json.load(f)

                existing["sessies"]      = sessies
                existing["generated_at"] = datetime.now(timezone.utc).isoformat()

                with fp.open("w", encoding="utf-8") as f:
                    json.dump(existing, f, ensure_ascii=False, indent=2)

                # Bij het standaardsseizoen ook roster_export.json bijwerken
                if slug == _DEFAULT_SLUG:
                    with ROSTER.open("w", encoding="utf-8") as f:
                        json.dump(existing, f, ensure_ascii=False, indent=2)

                # Excel bijwerken via planner --from-json
                if PLANNER_PY.exists():
                    subprocess.run(
                        [sys.executable, str(PLANNER_PY),
                         "--file", str(EXCEL),
                         "--from-json", str(fp)],
                        timeout=30,
                        cwd=str(PLANNER_PY.parent),
                        check=False,
                    )

                self._json({"ok": True, "count": len(sessies)})
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        elif re.match(r"^/api/seasons/[\w]+/delete$", self.path):
            try:
                slug = self.path.split("/")[3]
                fp   = SEASONS_DIR / f"{slug}.json"
                if not fp.exists():
                    self._json({"ok": False, "detail": f"Seizoen '{slug}' bestaat niet"})
                    return
                fp.unlink()
                self._json({"ok": True, "slug": slug})
            except Exception as e:
                import traceback; traceback.print_exc()
                self._json({"ok": False, "detail": str(e)})

        else:
            self.send_error(404)

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def _file(self, path, ct):
        try:
            data = Path(path).read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", ct)
            self._cors()
            self.end_headers()
            self.wfile.write(data)
        except FileNotFoundError:
            self.send_error(404)

    def _json(self, obj):
        data = json.dumps(obj, ensure_ascii=False).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self._cors()
        self.end_headers()
        self.wfile.write(data)

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def log_message(self, fmt, *args):
        pass


if __name__ == "__main__":
    import os
    port   = int(os.environ.get("PORT", 5051))
    host   = "0.0.0.0"
    server = HTTPServer((host, port), Handler)
    url    = f"http://localhost:{port}"
    print(f"JEKA Capaciteitsmanager → {url}")
    print("Druk Ctrl+C om te stoppen.")
    if not os.environ.get("PORT"):
        Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer gestopt.")
