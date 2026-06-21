"""
Eenmalig update-script: DATA + LOGICA sheet aanpassen o.b.v. documentwensen.
Uitvoeren met: python3 "update_excel.py"
"""
import re
import openpyxl

BESTAND = "Trainingschema planner v2.xlsx"

wb = openpyxl.load_workbook(BESTAND)

# ── DATA sheet ──────────────────────────────────────────────────────────────
ws_data = wb["DATA"]
headers = {cell.value: cell.column for cell in ws_data[1] if cell.value}

col_id   = headers.get("TeamID")
col_naam = headers.get("Team_naam")
col_act  = headers.get("Actief")
col_bijz = headers.get("Bijzonderheden")
col_cat  = headers.get("Categorie")

# Dagvoorkeuren per team-id:  {team_id: dag_code}
DAG_VOORKEUREN = {
    # Senioren-heren dagvoorkeur
    "Heren-3":  "WO",
    "Heren-4":  "DO",
    "Heren-5":  "DO",
    "Heren-6":  "WO",
    "Heren-7":  "WO",
    "Heren-8":  "DO",
    "Heren-9":  "WO",
    "Heren-10": "DO",
    "Heren-11": "DO",
    "Heren-12": "DI",
    "Heren-13": "DO",
    "Heren-14": "DO",
    "Heren-15": "MA",
    "Heren-16": "WO",
    "Heren-18": "DI",
    "Heren-19": "DI",
    # Dames
    "Dames-2":  "WO",
    # Senioren-selectie
    "Heren-1":  "DI",
    "Heren-2":  "DI",
    "Dames-1":  "DI",
}

ACTIEF_WIJZIGINGEN = {
    "Heren-17": False,
    "Heren-19": True,
}

CATEGORIE_WIJZIGINGEN = {
    "Dames-2": "Senioren",
}

VELD_REMAP = {
    "veld:3": "veld:1",
}

def bijz_set_dag(bijz_str, dag):
    """Vervang of voeg dag:XX toe in bijzonderheden."""
    bijz_str = bijz_str or ""
    if re.search(r"\bdag:[a-z]{2}\b", bijz_str):
        return re.sub(r"\bdag:[a-z]{2}\b", f"dag:{dag.lower()}", bijz_str)
    return (bijz_str.strip() + f" dag:{dag.lower()}").strip()

def bijz_remap_veld(bijz_str):
    """Herschrijf veld:3 naar veld:1 in bijzonderheden."""
    if not bijz_str:
        return bijz_str
    for oud, nieuw in VELD_REMAP.items():
        bijz_str = bijz_str.replace(oud, nieuw)
    return bijz_str

changes = []

for row in ws_data.iter_rows(min_row=2):
    if not row[0].value and not (col_naam and row[col_naam-1].value):
        continue
    team_id = row[col_id-1].value if col_id else None
    if not team_id:
        continue
    team_id = str(team_id).strip()

    # Bijzonderheden: dag-voorkeur
    if team_id in DAG_VOORKEUREN and col_bijz:
        cel = row[col_bijz-1]
        oude = str(cel.value or "")
        nieuwe = bijz_set_dag(oude, DAG_VOORKEUREN[team_id])
        if oude != nieuwe:
            cel.value = nieuwe
            changes.append(f"  {team_id} bijz: '{oude}' → '{nieuwe}'")

    # Bijzonderheden: veld:3 → veld:1
    if col_bijz:
        cel = row[col_bijz-1]
        oude = str(cel.value or "")
        nieuwe = bijz_remap_veld(oude)
        if oude != nieuwe:
            cel.value = nieuwe
            if f"  {team_id} bijz:" not in "".join(changes):
                changes.append(f"  {team_id} veld bijz: '{oude}' → '{nieuwe}'")

    # Actief-status
    if team_id in ACTIEF_WIJZIGINGEN and col_act:
        cel = row[col_act-1]
        oude = cel.value
        nieuw = ACTIEF_WIJZIGINGEN[team_id]
        if bool(oude) != nieuw:
            cel.value = nieuw
            changes.append(f"  {team_id} actief: {oude} → {nieuw}")

    # Categorie
    if team_id in CATEGORIE_WIJZIGINGEN and col_cat:
        cel = row[col_cat-1]
        oude = str(cel.value or "")
        nieuw = CATEGORIE_WIJZIGINGEN[team_id]
        if oude != nieuw:
            cel.value = nieuw
            changes.append(f"  {team_id} categorie: '{oude}' → '{nieuw}'")

# ── LOGICA sheet ─────────────────────────────────────────────────────────────
ws_logica = wb["LOGICA"]

# Kolom-headers zoeken in de LOGICA tabel
logica_headers = {}
in_tabel = False
header_row_idx = None

for i, row in enumerate(ws_logica.iter_rows(), start=1):
    if row[0].value and "tblCategorieRegels" in str(row[0].value):
        in_tabel = True
        continue
    if in_tabel and not logica_headers:
        logica_headers = {str(cell.value).strip(): cell.column
                          for cell in row if cell.value}
        header_row_idx = i
        continue
    if in_tabel and row[0].value is None:
        break

col_logica_cat  = logica_headers.get("Categorie")
col_logica_van  = logica_headers.get("Tijd_van")
col_logica_tot  = logica_headers.get("Tijd_tot")

# Tijdwijzigingen: {categorie: (nieuw_tijd_van, nieuw_tijd_tot)}
LOGICA_TIJD = {
    "Senioren":           ("20:30", None),
    "Senioren-selectie":  ("19:00", "21:00"),
}

if col_logica_cat:
    for row in ws_logica.iter_rows(min_row=header_row_idx+1 if header_row_idx else 2):
        cat_cel = row[col_logica_cat-1]
        cat = str(cat_cel.value or "").strip()
        if cat in LOGICA_TIJD:
            nieuw_van, nieuw_tot = LOGICA_TIJD[cat]
            if col_logica_van and nieuw_van:
                cel = row[col_logica_van-1]
                oude = cel.value
                cel.value = nieuw_van
                changes.append(f"  LOGICA {cat} Tijd_van: {oude} → {nieuw_van}")
            if col_logica_tot and nieuw_tot:
                cel = row[col_logica_tot-1]
                oude = cel.value
                cel.value = nieuw_tot
                changes.append(f"  LOGICA {cat} Tijd_tot: {oude} → {nieuw_tot}")

# ── Opslaan ──────────────────────────────────────────────────────────────────
wb.save(BESTAND)
print(f"Opgeslagen: {BESTAND}")
print(f"{len(changes)} wijzigingen:")
for c in changes:
    print(c)
