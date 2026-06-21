"""
setup_excel_v2.py — Setup voor RKVV JEKA Trainingsplanner versie 2
Maakt een nieuw Excel-bestand aan (of repareert een bestaand) met:
  - DATA-blad (5 kolommen: TeamID, Team_naam, Categorie, Actief, Bijzonderheden)
  - LOGICA-blad (categorie-regels, ongewijzigd t.o.v. v1)
  - ROOSTER-blad (visueel rooster, ongewijzigd t.o.v. v1)
  - DATA_LEEG + ROOSTER_LEEG (templates)

Geen FORMULIER-blad — teams hoeven geen voorkeuren meer in te vullen.
"""

import os
import shutil
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE_V2 = "Trainingschema planner v2.xlsx"
EXCEL_FILE_V1 = "../Trainingschema planner.xlsx"

# Teams met vaste tijden/dagen of speciale planningsregels.
# Dit zijn GEEN gebruikersvoorkeuren, maar structurele beperkingen.
FIXED_TEAMS = {
    "G-1":            "vast tijdstip 19:00-20:30; dag:WO; veld:3",
    "G-2":            "vast tijdstip 19:00-20:30; dag:WO; veld:3",
    "G-3":            "vast tijdstip 16:00-17:00; dag:VR; veld:1; gebruik:1.0",
    "G-4":            "vast tijdstip 17:00-18:15; dag:VR; veld:1; duur:75; gebruik:1.0",
    "G-5":            "vast tijdstip 18:15-19:30; dag:VR; veld:1; duur:75",
    "Keeperstraining": "vast tijdstip 17:45-19:00; dag:DI; veld:3; gebruik:0; duur:75",
    "KT-avond":       "vast tijdstip 19:00-20:30; dag:DI; veld:3; gebruik:0; duur:90",
    "KT-vrijdag":     "vast tijdstip 18:30-19:30; dag:VR; veld:3; gebruik:0; duur:60",
    "4SKILLS":        "vast tijdstip 16:00-17:00; dag:WO; veld:3",
    "Vaders-1":       "bi-wekelijks",
    "JO13-2":         "schaduwtraining JO13-1",
    "JO13-3":         "schaduwtraining JO13-1",
}

CATEGORIE_REGELS = [
    ("JO7",        60, 0.25, 1, "onderbouw",  "16:00", "19:00"),
    ("JO8",        60, 0.25, 1, "onderbouw",  "16:00", "19:00"),
    ("JO9",        60, 0.50, 1, "onderbouw",  "16:00", "19:00"),
    ("JO10",       60, 0.50, 1, "onderbouw",  "16:00", "19:00"),
    ("JO11",       60, 0.50, 1, "onderbouw",  "16:00", "20:30"),
    ("JO12",       60, 0.50, 1, "middenbouw", "17:00", "21:00"),
    ("JO13",       60, 0.50, 2, "middenbouw", "17:00", "21:00"),
    ("JO14",       60, 0.50, 2, "middenbouw", "16:30", "21:00"),
    ("JO15",       60, 0.50, 2, "middenbouw", "16:30", "19:30"),
    ("MO13",       60, 0.50, 2, "middenbouw", "16:30", "21:00"),
    ("MO15",       60, 0.50, 2, "middenbouw", "16:30", "19:30"),
    ("JO17",       90, 0.50, 2, "bovenbouw",  "19:30", "21:00"),
    ("JO19",       90, 0.50, 2, "bovenbouw",  "19:30", "21:00"),
    ("MO17",       90, 0.50, 2, "bovenbouw",  "19:30", "21:00"),
    ("MO20",       90, 0.50, 2, "bovenbouw",  "19:30", "21:00"),
    ("Senioren",           90, 0.50, 2, "senioren",   "20:00", "22:30", 0.25),
    ("Senioren-selectie",  90, 0.50, 2, "senioren-selectie", "20:00", "22:30", 0.50),
    ("Bijzonder",          60, 0.50, 1, "bijzonder",  "16:00", "22:30"),
    ("Gehandicapt",60, 0.50, 1, "bijzonder",  "19:00", "20:30"),
]

KLEUR_HEADER    = "1F3864"
KLEUR_SUBHEADER = "2E75B6"
KLEUR_ONDERBOUW  = "AED6F1"
KLEUR_MIDDENBOUW = "A9DFBF"
KLEUR_BOVENBOUW  = "F9E79F"
KLEUR_SENIOREN   = "F1948A"
KLEUR_BIJZONDER  = "D7BDE2"
KLEUR_GRIJS      = "F2F2F2"


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def make_border(thick=False):
    style = "medium" if thick else "thin"
    side  = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)


def header_font(bold=True, white=False, size=11):
    color = "FFFFFF" if white else "000000"
    return Font(bold=bold, color=color, size=size)


# ---------------------------------------------------------------------------
# Stap 1: DATA-blad opbouwen (5 kolommen, geen voorkeuren)
# ---------------------------------------------------------------------------

def repair_data_sheet_v2(wb):
    """
    Leest bestaande teams uit het DATA-blad en schrijft een vereenvoudigd
    DATA-blad met alleen: TeamID, Team_naam, Categorie, Actief, Bijzonderheden.
    Bijzonderheden wordt alleen ingevuld voor teams in FIXED_TEAMS.
    """
    ws_old = wb["DATA"]
    old_headers = {cell.value: cell.column - 1 for cell in ws_old[1] if cell.value}

    # Lees bestaande teams (alleen basiskolommen)
    teams = []
    existing_names = set()
    for row in ws_old.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        team_id   = row[old_headers["TeamID"]]   if "TeamID"   in old_headers else None
        team_naam = row[old_headers["Team_naam"]] if "Team_naam" in old_headers else None
        categorie = row[old_headers["Categorie"]] if "Categorie" in old_headers else None
        actief    = row[old_headers["Actief"]]    if "Actief"   in old_headers else None

        if not team_naam:
            continue

        # Repareer "=FALSE" string
        if actief == "=FALSE":
            actief = False

        teams.append({
            "team_id":   team_id or team_naam,
            "team_naam": team_naam,
            "categorie": categorie or "",
            "actief":    actief,
        })
        existing_names.add(str(team_naam))

    # Voeg vaste teams toe die nog niet bestaan
    for fixed_naam in FIXED_TEAMS:
        if fixed_naam not in existing_names:
            teams.append({
                "team_id":   fixed_naam,
                "team_naam": fixed_naam,
                "categorie": "",
                "actief":    True,
            })
            print(f"    Nieuw team toegevoegd: {fixed_naam}")

    # Verwijder oud DATA-blad en maak nieuw met 5 kolommen
    del wb["DATA"]
    ws = wb.create_sheet("DATA", 0)

    new_headers = ["TeamID", "Team_naam", "Categorie", "Actief", "Bijzonderheden"]
    for col_idx, h in enumerate(new_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = make_fill(KLEUR_SUBHEADER)
        cell.font      = header_font(white=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for row_idx, team in enumerate(teams, 2):
        bijz = FIXED_TEAMS.get(str(team["team_naam"]), "")
        ws.cell(row=row_idx, column=1, value=team["team_id"])
        ws.cell(row=row_idx, column=2, value=team["team_naam"])
        ws.cell(row=row_idx, column=3, value=team["categorie"])
        ws.cell(row=row_idx, column=4, value=team["actief"])
        ws.cell(row=row_idx, column=5, value=bijz)

        fill = make_fill("EBF5FB") if row_idx % 2 == 0 else make_fill("FDFEFE")
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 58

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2E75B6"

    print(f"  DATA-blad opgebouwd: {len(teams)} teams, 5 kolommen (geen voorkeuren).")


# ---------------------------------------------------------------------------
# Stap 2: LOGICA-blad aanmaken (ongewijzigd t.o.v. v1)
# ---------------------------------------------------------------------------

def create_logica_sheet(wb):
    if "LOGICA" in wb.sheetnames:
        del wb["LOGICA"]
    ws = wb.create_sheet("LOGICA", 1)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    row = 1

    def tbl_header(r, text):
        ws.merge_cells(f"A{r}:G{r}")
        ws[f"A{r}"] = text
        ws[f"A{r}"].fill      = make_fill(KLEUR_HEADER)
        ws[f"A{r}"].font      = Font(bold=True, color="FFFFFF", size=12)
        ws[f"A{r}"].alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws.row_dimensions[r].height = 22

    def col_headers(r, cols):
        for i, h in enumerate(cols, 1):
            c = get_column_letter(i)
            ws[f"{c}{r}"] = h
            ws[f"{c}{r}"].fill      = make_fill(KLEUR_SUBHEADER)
            ws[f"{c}{r}"].font      = Font(bold=True, color="FFFFFF")
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18

    # tblCategorieRegels
    tbl_header(row, "tblCategorieRegels — Regels per leeftijdscategorie")
    row += 1
    col_headers(row, ["Categorie", "Duur_min", "Veldgebruik", "Sessies_week",
                      "Prioriteit", "Tijd_van", "Tijd_tot", "Veldgebruik_2"])
    row += 1

    for regel in CATEGORIE_REGELS:
        cat, duur, veld, ses, prio, t_van, t_tot = regel[:7]
        veld2 = regel[7] if len(regel) > 7 else ""
        ws[f"A{row}"] = cat
        ws[f"B{row}"] = duur
        ws[f"C{row}"] = veld
        ws[f"D{row}"] = ses
        ws[f"E{row}"] = prio
        ws[f"F{row}"] = t_van
        ws[f"G{row}"] = t_tot
        ws[f"H{row}"] = veld2 if veld2 != "" else None
        fill_color = KLEUR_GRIJS if row % 2 == 0 else "FFFFFF"
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"].fill      = make_fill(fill_color)
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
            ws[f"{col}{row}"].border    = make_border()
        ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
        row += 1

    row += 1

    # tblVelden
    tbl_header(row, "tblVelden — Beschikbare velden")
    row += 1
    col_headers(row, ["Veld", "Capaciteit", "", "", "", "", ""])
    row += 1
    for veld_naam, cap in [("Veld 1", 1.0), ("Veld 2", 1.0), ("Veld 3", 1.0)]:
        ws[f"A{row}"] = veld_naam
        ws[f"B{row}"] = cap
        for col in ["A", "B"]:
            ws[f"{col}{row}"].border    = make_border()
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
        row += 1

    row += 1

    # tblTijdsloten
    tbl_header(row, "tblTijdsloten — Geldige starttijden (16:00–22:30, stap 15 min)")
    row += 1
    col_headers(row, ["Tijdslot", "", "", "", "", "", ""])
    row += 1

    t     = datetime.datetime.combine(datetime.date.today(), datetime.time(16, 0))
    t_end = datetime.datetime.combine(datetime.date.today(), datetime.time(22, 30))
    while t.time() <= t_end.time():
        ws[f"A{row}"] = t.strftime("%H:%M")
        ws[f"A{row}"].border    = make_border()
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        t += datetime.timedelta(minutes=15)
        row += 1

    row += 1

    # tblPrioriteitsgroepen
    tbl_header(row, "tblPrioriteitsgroepen — Planningsvolgorde (automatisch)")
    row += 1
    col_headers(row, ["Volgorde", "Groep", "Categorieen", "Tijd_van", "Tijd_tot", "", ""])
    row += 1
    prioriteiten = [
        (1, "bijzonder",  "G-teams, Vaders, Keeperstraining, 4SKILLS", "16:00", "22:30"),
        (2, "onderbouw",  "JO7 t/m JO11",                              "16:00", "20:30"),
        (3, "middenbouw", "JO12 t/m JO15, MO13, MO15",                 "17:00", "21:00"),
        (4, "bovenbouw",  "JO17 t/m JO19, MO17, MO20",                 "18:00", "21:00"),
        (5, "senioren",   "Heren, Dames",                               "20:30", "22:30"),
    ]
    for volgorde, groep, cats, t_van, t_tot in prioriteiten:
        ws[f"A{row}"] = volgorde
        ws[f"B{row}"] = groep
        ws[f"C{row}"] = cats
        ws[f"D{row}"] = t_van
        ws[f"E{row}"] = t_tot
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].border    = make_border()
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].alignment = Alignment(horizontal="left", indent=1)
        ws[f"C{row}"].alignment = Alignment(horizontal="left", indent=1)
        row += 1

    print("  LOGICA-blad aangemaakt.")


# ---------------------------------------------------------------------------
# Stap 3: ROOSTER-blad aanmaken (ongewijzigd t.o.v. v1)
# ---------------------------------------------------------------------------

def create_rooster_sheet(wb):
    if "ROOSTER" in wb.sheetnames:
        del wb["ROOSTER"]
    ws = wb.create_sheet("ROOSTER", 2)

    DAGEN_LABELS = ["Maandag", "Dinsdag", "Woensdag", "Donderdag", "Vrijdag"]
    SUBVELDEN    = ["Veld 1A", "Veld 1B", "Veld 2A", "Veld 2B", "Veld 3A", "Veld 3B"]

    tijdblokken = []
    t     = datetime.datetime.combine(datetime.date.today(), datetime.time(16, 0))
    t_end = datetime.datetime.combine(datetime.date.today(), datetime.time(22, 30))
    while t.time() <= t_end.time():
        tijdblokken.append(t.strftime("%H:%M"))
        t += datetime.timedelta(minutes=15)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for i in range(5):
        ws.column_dimensions[get_column_letter(3 + i)].width = 22

    total_cols = 2 + 5

    # Titel
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = "RKVV JEKA — Trainingsrooster v2 (volledig automatisch gegenereerd)"
    ws["A1"].fill      = make_fill(KLEUR_HEADER)
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Genereer-knop placeholder
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    ws["A2"] = "► GENEREER ROOSTER  (koppel VBA-macro 'GenereerRooster')"
    ws["A2"].fill      = make_fill("E74C3C")
    ws["A2"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # Kleurlegenda
    ws["A3"] = "Legenda:"
    ws["A3"].font = Font(bold=True)
    ws.row_dimensions[3].height = 18

    legenda = [
        ("Onderbouw (JO7-11)", KLEUR_ONDERBOUW, "C3"),
        ("Middenbouw (JO12-15)", KLEUR_MIDDENBOUW, "D3"),
        ("Bovenbouw (JO17-19)", KLEUR_BOVENBOUW, "E3"),
        ("Senioren", KLEUR_SENIOREN, "F3"),
        ("Bijzonder", KLEUR_BIJZONDER, "G3"),
    ]
    for tekst, kleur, cel in legenda:
        ws[cel] = tekst
        ws[cel].fill      = make_fill(kleur)
        ws[cel].font      = Font(bold=True, size=9)
        ws[cel].alignment = Alignment(horizontal="center", vertical="center")
        ws[cel].border    = make_border()

    # Dagheaders
    ws["A4"] = "Tijd"
    ws["B4"] = "SubVeld"
    for cell_ref in ["A4", "B4"]:
        ws[cell_ref].fill      = make_fill(KLEUR_SUBHEADER)
        ws[cell_ref].font      = Font(bold=True, color="FFFFFF")
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    for i, dag in enumerate(DAGEN_LABELS, 3):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}4"] = dag
        ws[f"{col_letter}4"].fill      = make_fill(KLEUR_SUBHEADER)
        ws[f"{col_letter}4"].font      = Font(bold=True, color="FFFFFF")
        ws[f"{col_letter}4"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_letter}4"].border    = make_border()

    # Grid
    current_row = 5
    for tijdblok in tijdblokken:
        merge_start = current_row
        merge_end   = current_row + len(SUBVELDEN) - 1

        ws.merge_cells(f"A{merge_start}:A{merge_end}")
        ws[f"A{merge_start}"] = tijdblok
        ws[f"A{merge_start}"].fill      = make_fill(KLEUR_GRIJS)
        ws[f"A{merge_start}"].font      = Font(bold=True, size=9)
        ws[f"A{merge_start}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{merge_start}"].border    = make_border()

        for sv_idx, subveld in enumerate(SUBVELDEN):
            r = current_row + sv_idx
            ws.row_dimensions[r].height = 16

            ws[f"B{r}"] = subveld
            sv_kleur    = "FDFEFE" if sv_idx % 2 == 0 else KLEUR_GRIJS
            ws[f"B{r}"].fill      = make_fill(sv_kleur)
            ws[f"B{r}"].font      = Font(size=8, italic=True)
            ws[f"B{r}"].alignment = Alignment(horizontal="left", indent=1, vertical="center")
            ws[f"B{r}"].border    = make_border()

            for dag_idx in range(5):
                col_letter = get_column_letter(3 + dag_idx)
                cell = ws[f"{col_letter}{r}"]
                cell.fill      = make_fill("FDFEFE")
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = make_border()
                cell.font      = Font(size=8)

        current_row += len(SUBVELDEN)

    # Sessietabel
    sessie_start_row = current_row + 2
    ws.merge_cells(f"A{sessie_start_row}:{get_column_letter(total_cols)}{sessie_start_row}")
    ws[f"A{sessie_start_row}"] = "SESSIETABEL"
    ws[f"A{sessie_start_row}"].fill      = make_fill(KLEUR_HEADER)
    ws[f"A{sessie_start_row}"].font      = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{sessie_start_row}"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[sessie_start_row].height = 20

    sessie_header_row = sessie_start_row + 1
    sessie_cols = ["TeamID", "Dag", "Starttijd", "Eindtijd", "Veld", "SubVeld", "Status"]
    for i, h in enumerate(sessie_cols, 1):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}{sessie_header_row}"] = h
        ws[f"{col_letter}{sessie_header_row}"].fill      = make_fill(KLEUR_SUBHEADER)
        ws[f"{col_letter}{sessie_header_row}"].font      = Font(bold=True, color="FFFFFF")
        ws[f"{col_letter}{sessie_header_row}"].alignment = Alignment(horizontal="center")
        ws[f"{col_letter}{sessie_header_row}"].border    = make_border()
    ws.row_dimensions[sessie_header_row].height = 18

    # Uitzonderingsrapport
    uitzondering_start_row = sessie_start_row + 300
    ws.merge_cells(f"A{uitzondering_start_row}:{get_column_letter(total_cols)}{uitzondering_start_row}")
    ws[f"A{uitzondering_start_row}"] = "UITZONDERINGSRAPPORT — Niet-ingeplande teams"
    ws[f"A{uitzondering_start_row}"].fill      = make_fill("C0392B")
    ws[f"A{uitzondering_start_row}"].font      = Font(bold=True, color="FFFFFF", size=11)
    ws[f"A{uitzondering_start_row}"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[uitzondering_start_row].height = 20

    uitzondering_header_row = uitzondering_start_row + 1
    for i, h in enumerate(["TeamID", "Reden"], 1):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}{uitzondering_header_row}"] = h
        ws[f"{col_letter}{uitzondering_header_row}"].fill      = make_fill(KLEUR_SUBHEADER)
        ws[f"{col_letter}{uitzondering_header_row}"].font      = Font(bold=True, color="FFFFFF")
        ws[f"{col_letter}{uitzondering_header_row}"].alignment = Alignment(horizontal="center")
        ws[f"{col_letter}{uitzondering_header_row}"].border    = make_border()

    # Metadata voor planner_v2.py
    ws["J1"] = sessie_header_row + 1
    ws["J2"] = uitzondering_header_row + 1
    ws["J1"].font = Font(color="FFFFFF", size=6)
    ws["J2"].font = Font(color="FFFFFF", size=6)

    ws.freeze_panes = "C5"
    ws.sheet_properties.tabColor = "2E75B6"

    print(f"  ROOSTER-blad aangemaakt ({len(tijdblokken)} tijdblokken × 6 subvelden).")


# ---------------------------------------------------------------------------
# Stap 4: DATA_LEEG-blad aanmaken
# ---------------------------------------------------------------------------

def create_leeg_sheet_v2(wb):
    for name in ("LEEG", "DATA_LEEG"):
        if name in wb.sheetnames:
            del wb[name]
    ws_leeg = wb.create_sheet("DATA_LEEG")
    ws_data = wb["DATA"]

    headers = [cell.value for cell in ws_data[1] if cell.value]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_leeg.cell(row=1, column=col_idx, value=header)
        cell.fill      = make_fill(KLEUR_SUBHEADER)
        cell.font      = header_font(white=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    leeg_row = 2
    for data_row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        team_naam = data_row[1].value  # kolom B
        if not team_naam:
            continue
        for col_idx, src_cell in enumerate(data_row[:len(headers)], 1):
            ws_leeg.cell(row=leeg_row, column=col_idx, value=src_cell.value)
        fill = make_fill("EBF5FB") if leeg_row % 2 == 0 else make_fill("FDFEFE")
        for col_idx in range(1, len(headers) + 1):
            ws_leeg.cell(row=leeg_row, column=col_idx).fill = fill
        leeg_row += 1

    ws_leeg.column_dimensions["A"].width = 14
    ws_leeg.column_dimensions["B"].width = 22
    ws_leeg.column_dimensions["C"].width = 14
    ws_leeg.column_dimensions["D"].width = 8
    ws_leeg.column_dimensions["E"].width = 58

    ws_leeg.sheet_properties.tabColor = "BDC3C7"
    ws_leeg.freeze_panes = "A2"

    print(f"  DATA_LEEG-blad aangemaakt: {leeg_row - 2} teams gekopieerd.")


# ---------------------------------------------------------------------------
# Stap 5: ROOSTER_LEEG-blad aanmaken
# ---------------------------------------------------------------------------

def create_leeg_rooster_sheet(wb):
    if "ROOSTER_LEEG" in wb.sheetnames:
        del wb["ROOSTER_LEEG"]

    ws_source = wb["ROOSTER"]
    ws_leeg   = wb.copy_worksheet(ws_source)
    ws_leeg.title = "ROOSTER_LEEG"

    ws_leeg["A1"] = "RKVV JEKA v2 — Leeg rooster-sjabloon"
    ws_leeg["A1"].fill      = make_fill("27AE60")
    ws_leeg["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws_leeg["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_leeg["A2"] = "Leeg sjabloon — genereer via knop of planner_v2.py --rooster-sheet ROOSTER_LEEG"
    ws_leeg["A2"].fill      = make_fill("1E8449")
    ws_leeg["A2"].font      = Font(bold=True, color="FFFFFF", size=10)
    ws_leeg["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws_leeg.sheet_properties.tabColor = "27AE60"

    print("  ROOSTER_LEEG-blad aangemaakt.")


# ---------------------------------------------------------------------------
# Hoofdprogramma
# ---------------------------------------------------------------------------

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_pad  = os.path.join(script_dir, EXCEL_FILE_V2)

    print(f"Setup v2 gestart voor: {excel_pad}\n")

    # Maak Excel-bestand als het nog niet bestaat
    if not os.path.exists(excel_pad):
        v1_pad = os.path.join(script_dir, EXCEL_FILE_V1)
        if os.path.exists(v1_pad):
            shutil.copy2(v1_pad, excel_pad)
            print(f"  Gekopieerd van: {v1_pad}")
        else:
            # Maak een minimaal nieuw bestand
            wb_new = openpyxl.Workbook()
            ws     = wb_new.active
            ws.title = "DATA"
            ws.append(["TeamID", "Team_naam", "Categorie", "Actief", "Bijzonderheden"])
            wb_new.save(excel_pad)
            print("  Nieuw leeg Excel-bestand aangemaakt (geen v1 gevonden).")

    wb = openpyxl.load_workbook(excel_pad)

    # Verwijder FORMULIER-blad als het nog aanwezig is (van v1)
    if "FORMULIER" in wb.sheetnames:
        del wb["FORMULIER"]
        print("  FORMULIER-blad verwijderd (niet nodig in v2).")

    print("1. DATA-blad opbouwen (5 kolommen, geen voorkeuren)...")
    repair_data_sheet_v2(wb)

    print("2. LOGICA-blad aanmaken...")
    create_logica_sheet(wb)

    print("3. ROOSTER-blad aanmaken...")
    create_rooster_sheet(wb)

    print("4. DATA_LEEG-blad aanmaken...")
    create_leeg_sheet_v2(wb)

    print("5. ROOSTER_LEEG-blad aanmaken...")
    create_leeg_rooster_sheet(wb)

    # Bladervolgorde
    expected_order = ["DATA", "LOGICA", "ROOSTER", "DATA_LEEG", "ROOSTER_LEEG"]
    wb._sheets.sort(key=lambda s: expected_order.index(s.title) if s.title in expected_order else 99)

    wb.save(excel_pad)
    print(f"\nKlaar! Excel v2 opgeslagen: {excel_pad}")
    print(f"Bladen: {wb.sheetnames}")


if __name__ == "__main__":
    main()
